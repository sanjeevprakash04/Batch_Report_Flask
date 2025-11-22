import os
os.environ["GIO_USE_VFS"] = "local"
os.environ["GDK_BACKEND"] = "none"
os.environ["NO_AT_BRIDGE"] = "1"
os.environ["WEASYPRINT_GUI"] = "false"
import pandas as pd
import base64
from datetime import datetime
import pdfkit
from weasyprint import HTML, CSS

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import platform

# wkhtml_path = '/usr/local/bin/wkhtmltopdf' if platform.system() == 'Linux' else 'data_files/wkhtmltopdf/bin/wkhtmltopdf.exe'


# ==========================================================
# 🔹 Utility Functions
# ==========================================================
def check(set_wg, act_wg, tol):
    if set_wg - tol <= act_wg <= set_wg + tol:
        return "Good"
    return "Up" if act_wg > set_wg + tol else "Down"

def difference(set_wg, act_wg):
    return round(act_wg - set_wg, 2)

def encode_logo(logo_path):
    """Converts logo to Base64 string for embedding."""
    with open(logo_path, "rb") as logo_file:
        return base64.b64encode(logo_file.read()).decode('utf-8')


# ==========================================================
# 🔹 PDF Report Generator
# ==========================================================
# def generate_pdf_report(df_pivot, df_string, batch_no):
#     # logo_base64 = encode_logo("data_files/logo.png")
#     logo_path = os.path.join("/app", "data_files", "logo.png")
#     logo_base64 = encode_logo(logo_path)

#     get_value = lambda k: df_string.loc[df_string["Name"] == k, "Value"].iloc[0] if not df_string[df_string["Name"] == k].empty else "N/A"
    
#     details = {
#         "printed_date": datetime.now().strftime("%d-%m-%Y %H:%M"),
#         "plant_name": get_value("Plant Name"),
#         "recipe_name": get_value("Recipe Name"),
#         "start_time": get_value("Start Date Time"),
#         "end_time": get_value("End Date Time"),
#         "batch_no": batch_no
#     }

#     try:
#         st = datetime.strptime(details["start_time"], '%Y-%m-%d %H:%M:%S')
#         ed = datetime.strptime(details["end_time"], '%Y-%m-%d %H:%M:%S')
#         details["time_taken"] = str(ed - st)
#     except Exception:
#         details["time_taken"] = "N/A"

#     details.update({
#         "total_set_weight": df_pivot["SetWeight"].sum(),
#         "total_actual_weight": round(df_pivot["ActualWeight"].sum(), 2),
#     })

#     html = generate_html_report(df_pivot, logo_base64, details)

#     pdf_bytes = pdfkit.from_string(
#         html,
#         False,
#         configuration=pdfkit.configuration(wkhtmltopdf=wkhtml_path),
#         options={
#             "enable-local-file-access": "",
#             "margin-top": "12mm",
#             "margin-bottom": "12mm",
#             "margin-left": "12mm",
#             "margin-right": "12mm",
#             "encoding": "UTF-8"
#         }
#     )
#     return pdf_bytes

def generate_pdf_report(df_pivot, df_string, batch_no):
    # Load logo
    logo_base64 = encode_logo("data_files/logo.png")

    get_value = lambda k: df_string.loc[df_string["Name"] == k, "Value"].iloc[0] \
        if not df_string[df_string["Name"] == k].empty else "N/A"

    details = {
        "printed_date": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "plant_name": get_value("Plant Name"),
        "recipe_name": get_value("Recipe Name"),
        "start_time": get_value("Start Date Time"),
        "end_time": get_value("End Date Time"),
        "batch_no": batch_no
    }

    # Time difference
    try:
        st = datetime.strptime(details["start_time"], '%Y-%m-%d %H:%M:%S')
        ed = datetime.strptime(details["end_time"], '%Y-%m-%d %H:%M:%S')
        details["time_taken"] = str(ed - st)
    except:
        details["time_taken"] = "N/A"

    details.update({
        "total_set_weight": df_pivot["SetWeight"].sum(),
        "total_actual_weight": round(df_pivot["ActualWeight"].sum(), 2),
    })

    # Generate HTML
    html = generate_html_report(df_pivot, logo_base64, details)

    # 🔥 Perfect scaling + A4 + margins (matches 2nd screenshot)
    pdf_bytes = HTML(string=html).write_pdf(
        stylesheets=[CSS(string="""
            @page {
                size: A4;
                margin: 6mm;
            }
            body {
                transform: scale(0.86);
                transform-origin: top center;
            }
        """)]
    )

    return pdf_bytes


# ==========================================================
# 🔹 Excel Report Generator
# ==========================================================
def generate_excel_report(df_pivot, df_string, batch_no):
    get_value = lambda k: df_string.loc[df_string["Name"] == k, "Value"].iloc[0] if not df_string[df_string["Name"] == k].empty else "N/A"
    details = {
        "printed_date": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "plant_name": get_value("Plant Name"),
        "recipe_name": get_value("Recipe Name"),
        "start_time": get_value("Start Date Time"),
        "end_time": get_value("End Date Time"),
        "batch_no": batch_no
    }

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Report"

    ws.append(["BATCH REPORT"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["Printed Date:", details["printed_date"]])
    ws.append([])

    headers = [
        ("Plant Name", details["plant_name"]),
        ("Recipe Name", details["recipe_name"]),
        ("Batch No", details["batch_no"]),
        ("Start Time", details["start_time"]),
        ("End Time", details["end_time"]),
        ("Total Set Weight (Kg)", df_pivot["SetWeight"].sum()),
        ("Total Actual Weight (Kg)", round(df_pivot["ActualWeight"].sum(), 2))
    ]
    for i in range(0, len(headers), 2):
        row = [headers[i][0], headers[i][1]]
        if i + 1 < len(headers):
            row += [headers[i+1][0], headers[i+1][1]]
        ws.append(row)
    ws.append([])

    table_headers = ["Silo No", "Material Name", "Set Weight", "Actual Weight", "Difference", "Tolerance"]
    ws.append(table_headers)

    header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    for col in range(1, len(table_headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    for _, row in df_pivot.iterrows():
        ws.append([
            row.get("SiloNo", ""),
            row.get("MaterialName", ""),
            row.get("SetWeight", ""),
            row.get("ActualWeight", ""),
            row.get("Difference", ""),
            row.get("Tolerance", "")
        ])

    wb.save(output)
    output.seek(0)
    return output.read()


# ==========================================================
# 🔹 HTML Template for PDF (Upgraded Alignment)
# ==========================================================
def generate_html_report(df, logo_base64, details):
    data_rows = "".join(f'''
        <tr>
            <td>{row["SiloNo"]}</td>
            <td>{row["MaterialName"]}</td>
            <td>{row["SetWeight"]}</td>
            <td>{row["ActualWeight"]}</td>
            <td>{row["Difference"]}</td>
            <td>{row["Tolerance"]}</td>
        </tr>
    ''' for _, row in df.iterrows())

    html_template = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Batch Report</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap');
            @page {{ size: A4; margin: 20mm; }}
            body {{ font-family: 'Roboto', sans-serif; margin: 20px; font-size: 12pt; }}
            h1 {{ text-align: center; margin-bottom: 20px; }}
            .header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; }}
            .printed-date {{ text-align: right; font-weight: bold; font-size: 12pt; }}
            .container {{ margin-bottom: 20px; }}
            .info-section table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            .info-section td {{ padding: 8px; text-align: left; border: none; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            td, th {{ padding: 10px; text-align: center; border: 1px solid #ddd; font-size: 11pt; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .footer {{ text-align: center; margin-top: 20px; font-weight: bold; }}
            .small-footer {{ text-align: center; font-size: 10px; margin-top: 10px; }}
        </style>
    </head>
    <body>
        <h1>BATCH REPORT</h1>
        <div class="header">
            <img src="data:image/png;base64,{logo_base64}" alt="Logo" style="width: 100px;">
            <div class="printed-date">Printed Date: {details['printed_date']}</div>
        </div>
        <div class="container">
            <div class="info-section">
                <table>
                    <tr><td><b>Recipe Name:</b></td><td>{details['recipe_name']}</td><td><b>Time Taken:</b></td><td>{details['time_taken']}</td></tr>
                    <tr><td><b>Batch No:</b></td><td>{details['batch_no']}</td><td><b>Total Set Weight:</b></td><td>{details['total_set_weight']} Kg</td></tr>
                    <tr><td><b>Start Time:</b></td><td>{details['start_time']}</td><td><b>Total Actual Weight:</b></td><td>{details['total_actual_weight']} Kg</td></tr>
                    <tr><td><b>End Time:</b></td><td>{details['end_time']}</td></tr>
                </table>
            </div>
        </div>
        <h2>Data Table</h2>
        <table>
            <tr>
                <th>Silo No</th>
                <th>Material Name</th>
                <th>Set Weight (Kg)</th>
                <th>Actual Weight (Kg)</th>
                <th>Difference (Kg)</th>
                <th>Tolerance (Kg)</th>
            </tr>
            {data_rows}
        </table>
        <div class="small-footer">This report is generated by Skew Reporting Software, developed by Prolite Automation.</div>
    </body>
    </html>
    '''
    return html_template
